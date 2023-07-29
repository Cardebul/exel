import openpyxl as op
from openpyxl.styles import Alignment, Border, Side, PatternFill, Font, Color
from datetime import datetime as dt, timedelta
"""Константы"""
warn = 'ошибка получения данных'
alignment = Alignment(horizontal='center', vertical='center')
fill = PatternFill(start_color="808080", end_color="808080", fill_type="solid")
side = Side(border_style="thin", color="000000")
bor = Border(left=side, top=side, right=side, bottom=side)
font = Font(color=Color(rgb='FFFFFF'))


def validator(data):
    """Принимает на вход словарь, делит словарь по этапам"""
    globalist = []
    step = data.get('Этап', warn)
    works = data.get('Работы', warn)
    count = data.get('Количество', warn)
    bug = data.get('Временной лаг', warn)
    full_step = {}
    perm_step = step[0]
    if len(set(step)) == 1:
        full_step[perm_step] = {
                'Работы': works,
                'Количество': count,  # fixed
                'Временной лаг': bug
            }
        globalist.append(full_step)
        return globalist
    obraz = {
        'Работы': [],
        'Количество': [],
        'Временной лаг': []
    }
    full_step[perm_step] = obraz
    lock = zip(step, works, count, bug)
    for i in lock:
        if i[0] != perm_step:
            globalist.append(full_step)
            full_step = {}
            perm_step = i[0]
            full_step[perm_step] = {
                        'Работы': [],
                        'Количество': [],
                        'Временной лаг': []
            }
        full_step[perm_step]['Работы'].append(i[1])
        full_step[perm_step]['Количество'].append(i[2])
        full_step[perm_step]['Временной лаг'].append(i[3])
    globalist.append(full_step)
    return globalist


def fil(sheet, row, fill, start=2, end=27):
    """Закраска"""
    for i in range(start, end+1):
        sheet.cell(row=row, column=i).fill = fill
    return sheet


def borders(sheet, row, side, start=2, end=27, mod=False):
    """Mod отвечает за границы для каждой ячеки, функция создает границы"""
    if mod:
        for i in range(start, end+1):
            sheet.cell(row=row, column=i).border = Border(top=side,
                                                          bottom=side,
                                                          left=side,
                                                          right=side)
        return sheet
    sheet.cell(row=row, column=start).border = Border(left=side, top=side,
                                                      bottom=side)
    sheet.cell(row=row, column=end).border = Border(top=side, bottom=side,
                                                    right=side)
    for i in range(start+1, end):
        sheet.cell(row=row, column=i).border = Border(top=side, bottom=side)
    return sheet


def create(sheet, data, row, date, cash, total_sum, hours, D_W):
    """
    Создание общей базы для каждего этапа.
    На вход принимает лист, данные, строку, дату, почасовую сумму,
    общую сумму и общее количество часов
    """
    D_or_W = 5 if D_W == 'неделям' else 1
    mul = 8*5 if D_W == 'неделям' else 8
    sheet.merge_cells(start_row=row, start_column=2, end_row=row, end_column=3)
    name, many = tuple(data.items())[0]
    font = Font(bold=True)
    sheet.cell(row=row, column=2).value = name
    sheet.cell(row=row, column=2).font = font
    for i in range(7, 28):
        sheet.cell(row=row, column=i).value = i - 6
        sheet.cell(row=row, column=i).font = font
    fill = PatternFill(start_color="a0deb0", end_color="a0deb0",
                       fill_type="solid")
    side = Side(border_style="thin", color="000000")
    fil(sheet, row, fill)
    borders(sheet, row, side)
    row += 1
    """Формат dt"""
    works = many.get('Работы')
    count = many.get('Количество')
    bug = many.get('Временной лаг')
    fin_data = zip(works, count, bug)
    col = 2
    for i in fin_data:
        val_1 = int(i[1])
        val_2 = int(i[2])
        sheet.merge_cells(start_row=row, start_column=2, end_row=row,
                          end_column=3)
        sheet.cell(row=row, column=col).value = i[0]
        time = dt.strptime(date, '%Y-%m-%d') + timedelta(days=val_2*D_or_W-1)
        sheet.cell(row=row, column=col+2).value = time.strftime('%d.%m.%Y')
        sheet.cell(row=row, column=col+3).value = val_1 * mul
        sheet.cell(row=row, column=col+3).alignment = alignment
        sheet.cell(row=row, column=col+4).value = val_1 * mul * cash
        total_sum += val_1 * mul * cash
        hours += val_1 * mul
        start = 6 + val_2
        end = start + val_1
        for j in range(start, end):
            sheet.cell(row=row, column=j).fill = fill
        borders(sheet, row, side, mod=True)
        row += 1
    return sheet, row, total_sum, hours


def ff(data1, data2):
    wb = op.Workbook()
    sheet = wb.active
    num_to_letter = op.utils.cell.get_column_letter
    val = [4, 35, 30, 10, 8, 13]
    val.extend([2.69 for _ in range(33)])
    for i in range(1, len(val)):
        sheet.column_dimensions[num_to_letter(i)].width = val[i-1]
    sheet.row_dimensions[5].height = 30
    sheet.cell(row=2, column=2).value = 'Дата начала'
    sheet.cell(row=3, column=2).value = 'Тариф'
    price = data1.get('Тариф', warn)
    price = price if isinstance(price, int) else int(price)
    sheet.cell(row=3, column=3).value = price  # TODO data1.get('Тариф')
    date = data1.get('Дата', warn)
    time = dt.strptime(date, '%Y-%m-%d').strftime('%d.%m.%Y')
    sheet.cell(row=2, column=3).value = time
    name = data1.get('Имя', warn)
    sheet.cell(row=2, column=7).value = name
    # границы
    sheet.cell(row=2, column=2).border = Border(left=side, top=side)
    sheet.cell(row=3, column=2).border = Border(left=side, bottom=side)
    sheet.cell(row=2, column=3).border = Border(right=side, top=side)
    sheet.cell(row=3, column=3).border = Border(bottom=side, right=side)
    sheet.cell(row=2, column=7).border = bor

    sheet.cell(row=2, column=2).fill = fill
    sheet.cell(row=3, column=2).fill = fill
    sheet.cell(row=2, column=3).fill = fill
    sheet.cell(row=3, column=3).fill = fill

    sheet.cell(row=2, column=2).font = font
    sheet.cell(row=3, column=2).font = font
    sheet.cell(row=2, column=3).font = font
    sheet.cell(row=3, column=3).font = font

    sheet.cell(row=2, column=2).alignment = alignment
    sheet.cell(row=3, column=2).alignment = alignment
    sheet.cell(row=2, column=3).alignment = alignment
    sheet.cell(row=3, column=3).alignment = alignment

    sheet.cell(row=2, column=7).alignment = alignment
    sheet.merge_cells(start_row=2, start_column=7, end_row=3, end_column=27)
    values = ['Задачи', 'Начало', 'Часы*', 'Стоимость']
    sheet.cell(row=5, column=2).value = values[0]
    sheet.cell(row=5, column=4).value = values[1]
    sheet.cell(row=5, column=5).value = values[2]
    sheet.cell(row=5, column=6).value = values[3]
    sheet.merge_cells(start_row=5, start_column=7, end_row=5, end_column=27)
    sheet.merge_cells(start_row=5, start_column=2, end_row=5, end_column=3)
    D_W = data1.get('Дни/Недели', warn)
    sheet.cell(row=5, column=7).value = f'Распределение объемов работ по {D_W}'

    borders(sheet, 5, side)
    # центрирование
    sheet.cell(row=5, column=2).alignment = alignment
    sheet.cell(row=5, column=4).alignment = alignment
    sheet.cell(row=5, column=5).alignment = alignment
    sheet.cell(row=5, column=6).alignment = alignment
    sheet.cell(row=5, column=7).alignment = alignment
    # закраска
    sheet.cell(row=5, column=2).fill = fill
    sheet.cell(row=5, column=4).fill = fill
    sheet.cell(row=5, column=5).fill = fill
    sheet.cell(row=5, column=6).fill = fill
    sheet.cell(row=5, column=7).fill = fill
    # шрифт
    sheet.cell(row=5, column=2).font = font
    sheet.cell(row=5, column=4).font = font
    sheet.cell(row=5, column=5).font = font
    sheet.cell(row=5, column=6).font = font
    sheet.cell(row=5, column=7).font = font

    globallist = validator(data2)
    row = 6
    total_sum = 0
    hours = 0
    for i in globallist:
        sheet, row, total_sum, hours = create(sheet, i, row, date,
                                              price, total_sum, hours, D_W)

    sheet.cell(row=row, column=2).value = 'Итого'
    sheet.cell(row=row, column=5).value = hours
    sheet.cell(row=row, column=6).value = total_sum
    new_side = Side(border_style="medium", color="000000")

    borders(sheet, row, new_side, end=6)
    name_file = f'{name}.xlsx'
    file_path = f'xlfiles/{name_file}'
    wb.save(file_path)
    return file_path
